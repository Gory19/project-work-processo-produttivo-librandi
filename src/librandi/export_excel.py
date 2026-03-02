from __future__ import annotations

from pathlib import Path

from .models import ScenarioSimulationResult


def export_result_to_excel(result: ScenarioSimulationResult, output_path: str | Path) -> Path:
    try:
        from openpyxl import Workbook
    except ModuleNotFoundError as exc:  # pragma: no cover - depends on local env
        raise SystemExit(
            "Per esportare in Excel serve 'openpyxl'. Installa con: pip install openpyxl"
        ) from exc

    path = Path(output_path).expanduser()
    if path.suffix.lower() != ".xlsx":
        path = path.with_suffix(".xlsx")
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Riepilogo"

    company_name, scenario_from_name = _split_company_and_scenario(result.scenario_name)
    scenario_selected = str(result.metadata.get("scenario_selected", scenario_from_name))
    weather_label = _weather_label(result.weather_condition)

    ws_summary.append(["Campo", "Valore"])
    ws_summary.append(["Azienda", company_name])
    ws_summary.append(["Scenario Meteo Selezionato", scenario_selected])
    ws_summary.append(["Sequenze Produttive Selezionate", ", ".join(result.selected_sequences)])
    ws_summary.append(
        [
            "Meteo Applicato",
            f"{weather_label} (x{result.weather_multiplier:.2f} sui tempi variabili)",
        ]
    )
    ws_summary.append(
        [
            "Tempo Massimo Lavorabile/Giorno",
            f"{result.max_minutes_per_day:.0f} min ({result.max_minutes_per_day / 60.0:.2f} h)",
        ]
    )
    ws_summary.append(
        [
            "Tempo Totale di Produzione del Processo",
            f"{result.total_minutes:.2f} min ({result.total_hours:.2f} h)",
        ]
    )
    ws_summary.append(
        ["Giorni Minimi Complessivi di Produzione", str(result.global_days_required)]
    )

    ws_products = wb.create_sheet("Prodotti")
    ws_products.append(
        [
            "Prodotto",
            "Quantita",
            "Unita",
            "Tempo Base (Meteo Buono) [min]",
            "Tempo Base (Meteo Applicato) [min]",
            "Tempo Totale Produzione [min]",
            "Tempo Totale Produzione [h]",
            "Giorni Necessari in Base alla Quantita del Prodotto",
            "Giorni Necessari in Base al Tempo Massimo Lavorabile al Giorno",
            "Giorni Minimi Totali di Produzione",
        ]
    )
    for item in result.per_product:
        ws_products.append(
            [
                item.product_label,
                item.quantity,
                result.quantity_unit,
                round(item.nominal_base_minutes, 2),
                round(item.base_minutes, 2),
                round(item.total_minutes, 2),
                round(item.total_minutes / 60.0, 2),
                item.days_by_quantity_capacity,
                item.days_by_time_capacity,
                item.min_days_required,
            ]
        )

    ws_sequences = wb.create_sheet("Sequenze")
    ws_sequences.append(["Prodotto", "Sequenza", "Tempo [min]", "Tempo [h]"])
    for item in result.per_product:
        for sequence_id, sequence_minutes in item.sequence_minutes.items():
            ws_sequences.append(
                [
                    item.product_label,
                    sequence_id.title(),
                    round(sequence_minutes, 2),
                    round(sequence_minutes / 60.0, 2),
                ]
            )

    _autofit_columns(ws_summary)
    _autofit_columns(ws_products)
    _autofit_columns(ws_sequences)

    wb.save(path)
    return path


def _split_company_and_scenario(full_name: str) -> tuple[str, str]:
    if " - " not in full_name:
        return full_name, "Non specificato"
    company, _, scenario = full_name.partition(" - ")
    return company.strip(), scenario.strip() or "Non specificato"


def _weather_label(condition: str) -> str:
    labels = {
        "good": "Meteo Buono",
        "bad": "Maltempo",
    }
    return labels.get(condition, condition)


def _autofit_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 60)
